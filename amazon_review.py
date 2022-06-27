import time

from selenium import webdriver
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.cell(row=1, column=1).value = "Customer Name"
ws.cell(row=1, column=2).value = "Review Date"
ws.cell(row=1, column=3).value = "Review Score"
ws.cell(row=1, column=4).value = "Review Title"
ws.cell(row=1, column=5).value = "Review content"
driver = webdriver.Chrome(executable_path="E:\\driver\\chromedriver.exe")
l = 2
m=2
n=2
o=2
p=2
for r in range(1, 845):
    print(r)
    driver.get(
        url="https://www.amazon.in/Redmi-Storage-Additional-Exchange-Included/product-reviews/B09T2XDXBN/ref=cm_cr_getr_d_paging_btm_prev_" + str(
            r) + "?ie=UTF8&pageNumber=" + str(r) + "&reviewerType=all_reviews")

    container = driver.find_element_by_id("cm_cr-review_list")
    for name in container.find_elements_by_class_name("a-profile-name"):
        print("Customer Name:\n", name.text)
        ws.cell(row=l, column=1).value = name.text
        l = l + 1

    wb.save("amazon.xlsx")
    for date in container.find_elements_by_class_name("review-date"):
        print("Review Date:\n", date.text)
        ws.cell(row=m, column=2).value = date.text
        m = m + 1

    wb.save("amazon.xlsx")
    for rating in container.find_elements_by_class_name("a-link-normal"):
        r2 = rating.get_attribute("title")
        print("Review Score", r2)
        ws.cell(row=n, column=3).value = r2
        n = n + 1

    wb.save("amazon.xlsx")
    for title in container.find_elements_by_class_name("review-title-content"):
        print("Review Title:\n", title.text)
        ws.cell(row=o, column=4).value = title.text
        o = o + 1

    wb.save("amazon.xlsx")
    for content in container.find_elements_by_class_name("review-text-content"):
        print("Review content:\n", content.text)
        ws.cell(row=p, column=5).value = content.text

        p = p + 1

    wb.save("amazon.xlsx")
# time.sleep(30)
driver.quit()

# import pandas as pd
#
# df = pd.read_excel("amazon.xlsx")
# new_df = df.dropna()
# print(new_df.to_string())
# df.to_csv("amazon1.csv")
