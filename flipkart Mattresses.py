
from selenium import webdriver
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
l = 2
ws.cell(row=1, column=1).value = "Product Name"
ws.cell(row=1, column=2).value = "Price"
ws.cell(row=1, column=3).value = "Size"
ws.cell(row=1, column=4).value = "Rating"
ws.cell(row=1, column=5).value = "Image Link"
ws.cell(row=1, column=6).value = "Product Link"
driver = webdriver.Chrome(executable_path="E:\\driver\\chromedriver.exe")

for r in range(1, 25):
    print(r)
    driver.get(url="https://www.flipkart.com/furniture/mattresses/pr?sid=wwe%2Crg9&otracker=nmenu_sub_Home+%26"
                   "+Furniture_0_Mattresses&page=" + str(r))
    for container in driver.find_elements_by_class_name("_4ddWXP"):
        for title in container.find_elements_by_class_name("s1Q9rs"):
            title.get_attribute('title')
            ws.cell(row=l,column=1).value = title.get_attribute('title')
            print("Product Name:", title.get_attribute('title'))
        for price in container.find_elements_by_class_name("_30jeq3"):
            print("Price", price.text)
            ws.cell(row=l,column=2).value = price.text
        for size in container.find_elements_by_class_name("_3Djpdu"):
            print("Size:", size.text)
            ws.cell(row=l,column=3).value = size.text
        for rating in container.find_elements_by_class_name("_2D5lwg"):
            print("Rating:", rating.text)
            ws.cell(row=l,column=4).value = rating.text
        for hyper_link in container.find_elements_by_class_name("_3exPp9"):
            link = hyper_link.get_attribute('src')
            print("Image Link",hyper_link.get_attribute('src'))
            ws.cell(row=l,column=5).value = link
        for hyper in container.find_elements_by_class_name("_2rpwqI"):
            link = hyper.get_attribute('href')
            print("Product Link",link)
            ws.cell(row=l,column=6).value = link
        l= l+1
    #
    #
    wb.save("Flipkart mattresses.xlsx")

driver.quit()
